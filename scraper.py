import os

import openpyxl
from selenium import webdriver
from selenium.webdriver import ChromeOptions, DesiredCapabilities
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from joblib import Parallel, delayed
import time


def scrap(negocio, lugar):
    url = "https://www.google.es/maps/search/"
    ruta = url + negocio + "+" + lugar.replace(" ", "+")
    opciones = ChromeOptions()
    opciones.add_argument("--incognito")
    opciones.add_argument("--no-sandbox")
    opciones.add_argument("--disable-dev-shm-usage")
    opciones.add_argument("--disable-gpu")
    opciones.add_argument("--ignore-certificate-errors")
    opciones.add_argument("--window-size=1920,1080")
    opciones.add_argument("--headless")
    navegador = webdriver.Chrome(opciones)
    #navegador.maximize_window()
    navegador.get(ruta)

    WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".VfPpkd-LgbsSe.VfPpkd-LgbsSe-OWXEXe-k8QpJ.VfPpkd-LgbsSe-OWXEXe-dgl2Hf.nCP5yc.AjY5Oe.DuMIQc.LQeN7.XWZjwc"))).click()

    previamente_contado = 0
    while True:
        lista_elementos = navegador.find_elements(By.CLASS_NAME, 'hfpxzc')

        action = ActionChains(navegador)
        action.move_to_element(lista_elementos[-1]).perform()

        time.sleep(5)
        scroll_origen = ScrollOrigin.from_element(lista_elementos[-1])
        action.scroll_from_origin(scroll_origen, 0, 1200).perform()

        time.sleep(20)
        action.scroll_from_origin(scroll_origen, 0, 250).perform()

        if len(lista_elementos) == previamente_contado:
            print(f'Total recopilado para {negocio} - {lugar}: {len(lista_elementos)}')
            break
        else:
            previamente_contado = len(lista_elementos)
            print(f'Actualmente recopilado para {negocio} - {lugar}: {len(lista_elementos)}')

    datos_negocio = []

    for x, elemento in enumerate(lista_elementos):
        time.sleep(2)
        if x == 0:
            accion_2 = ActionChains(navegador)
            accion_2.scroll_to_element(elemento).perform()
            accion_2.move_to_element(elemento).perform()

            time.sleep(2)
            elemento.click()
        else:
            entero = x - 1
            scroll_origen = ScrollOrigin.from_element(lista_elementos[entero])
            accion_2 = ActionChains(navegador)
            accion_2.scroll_to_element(elemento).perform()
            accion_2.scroll_from_origin(scroll_origen, 0, 190).perform()
            accion_2.scroll_to_element(elemento).perform()
            accion_2.move_to_element(elemento).perform()

            time.sleep(2)
            elemento.click()

        time.sleep(5)

        nombre_xpath = '//*[@id="QA0Szd"]/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]/div[2]/div/div[1]/div[1]/h1'
        direccion_xpath = '//button[@data-item-id="address"]//div[contains(@class, "fontBodyMedium")]'
        sitio_web_xpath = '//a[@data-item-id="authority"]//div[contains(@class, "fontBodyMedium")]'
        telefono_xpath = '//button[contains(@data-item-id, "phone:tel:")]//div[contains(@class, "fontBodyMedium")]'

        nombre = ''
        direccion = ''
        sitio_web = ''
        telefono = ''

        try:
            if len(navegador.find_element(By.XPATH, nombre_xpath).text) > 0:
                nombre = navegador.find_element(By.XPATH, nombre_xpath).text
        except NoSuchElementException:
            pass

        try:
            if len(navegador.find_element(By.XPATH, direccion_xpath).text) > 0:
                direccion = navegador.find_element(By.XPATH, direccion_xpath).text
        except NoSuchElementException:
            pass

        try:
            if len(navegador.find_element(By.XPATH, sitio_web_xpath).text) > 0:
                sitio_web = navegador.find_element(By.XPATH, sitio_web_xpath).text
        except NoSuchElementException:
            pass

        try:
            if len(navegador.find_element(By.XPATH, telefono_xpath).text) > 0:
                telefono = navegador.find_element(By.XPATH, telefono_xpath).text
        except NoSuchElementException:
            pass

        datos_negocio.append((nombre, direccion, sitio_web, telefono))

    wb = openpyxl.Workbook()
    hoja = wb.active

    for i, (nombre, direccion, sitio_web, telefono) in enumerate(datos_negocio, start=1):
        hoja.cell(row=i, column=1).value = f'{negocio} - {nombre} - {lugar}'
        hoja.cell(row=i, column=2).value = direccion
        hoja.cell(row=i, column=3).value = sitio_web
        hoja.cell(row=i, column=4).value = telefono

    wb.save(negocio + " " + lugar + ".xlsx")

    navegador.quit()


negocios = ["cafeteria", "chiringuito", "hotel", "pension", "hostal", "kebab", "hamburgueseria", "pizzeria", "restaurante", "bar-pub-discoteca"]

#lugares = sys.argv[1:]

lugares = ["malaga", "fuengirola"]

nucleos = os.cpu_count() // 2

for lugar in lugares:
    Parallel(n_jobs=nucleos)(delayed(scrap)(negocio, lugar) for negocio in negocios)
